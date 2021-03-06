import os
import pandas as pd
import openpyxl as op
from pandas import ExcelWriter
from pathlib import Path
from google_drive_api import *


class ExcelHandler():
    def __init__(self):
        self._sheet = pd.DataFrame({'Lám': [''],
                                    'N°Rta': [''],
                                    'Loc': [''],
                                    'DQ': [''],
                                    'Det': [''],
                                    'FQ': [''],
                                    '(2)': [''],
                                    'Cont': [''],
                                    'Pop': [''],
                                    'Pje Z': [''],
                                    'CCEE': [''],
                                    'respuesta': [''],
                                    'razon': ['']})
        self._sheet = self._sheet[['Lám', 'N°Rta', 'Loc', 'DQ', 'Det', 'FQ', '(2)', 'Cont', 'Pop', 'Pje Z', 'CCEE', 'respuesta', 'razon']]
        self._file_root = None

    def create_excel_sheet(self, nombre: str) -> str:
        self._file_root = str(self.get_project_root()) + \
            os.path.sep + 'files'+os.path.sep + nombre + '.xlsx'
        writer = ExcelWriter(self._file_root)
        self._sheet = self._sheet.to_excel(writer, 'Hoja de datos', index=False)
        writer.save()
        writer.close()

    def get_project_root(self) -> Path:
        return Path(__file__).parent.parent

    def upload_data(self, determinantes, contenidos, par, popular, dq, location, zscore, responses, reasons) -> str:
        wb = op.load_workbook(self._file_root)
        ws = wb.get_sheet_by_name('Hoja de datos')
        row = 2
        for lamina in range(3):
            rta = 0      
            for rta in range(len(responses[lamina])):
                ws['A' + str(row)]=lamina+1
                ws['B' + str(row)]=rta+1      # Número de respuesta 
                ws['C' + str(row)]=''.join([i for i in location[lamina][rta] if not i.isdigit()])   # Elimina los numeros de location
                ws['D' + str(row)]=dq[lamina][rta]
                ws['E' + str(row)]=(determinantes[lamina][rta])[:-1]
                ws['F' + str(row)]='?'
                ws['G' + str(row)]=par[lamina][rta]
                ws['H' + str(row)]=(contenidos[lamina][rta])[:-1]
                ws['I' + str(row)]=popular[lamina][rta]
                ws['J' + str(row)]=zscore[lamina][rta]
                ws['K' + str(row)]='?'
                ws['L' + str(row)]=responses[lamina][rta]
                ws['M' + str(row)]=reasons[lamina][rta]
                row += 1      
        
        wb.save(self._file_root)
        wb.close()         
        upload_file(self._file_root, "1EQ4h-Blfc3PqySXRvViSVrq2ZhCmq2rl")   

        planilla = pd.read_excel(self._file_root)
        print(planilla)
        os.remove(self._file_root)       
